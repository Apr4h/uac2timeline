"""
Export endpoint — generates an .xlsx workbook from tag-selected artifact rows.

GET /api/collections/{id}/export?tag_ids=1,2,3

Workbook structure
──────────────────
  Sheet 1  "Timeline"        — all tagged rows merged, sorted by timestamp,
                               with enrichment columns + notes
  Sheet N  per-artifact-type — all tagged rows for that type with full columns
                               + notes, one sheet per type that has results
"""
from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.app.database import get_db
from backend.app.models import Note, Tagging
from backend.app.api.timeline import (
    _enrich_event,
    _process_summary,
    _auth_summary,
    _cmdhistory_summary,
    _syslog_summary,
)
from uac_parser.models import (
    UACCollection,
    Authentication,
    CommandHistory,
    CronJob,
    File,
    NetworkConnection,
    Process,
    RcScript,
    SyslogEntry,
    SystemdService,
    User,
)

router = APIRouter()

# ── Constants ─────────────────────────────────────────────────────────────────

_MODEL_MAP = {
    "processes":  Process,
    "netconns":   NetworkConnection,
    "auth":       Authentication,
    "cmdhistory": CommandHistory,
    "users":      User,
    "files":      File,
    "cron":       CronJob,
    "services":   SystemdService,
    "rcscripts":  RcScript,
    "syslog":     SyslogEntry,
}

# Canonical export order for sheets
_TYPE_ORDER = [
    "processes", "auth", "cmdhistory", "netconns",
    "files", "cron", "services", "rcscripts", "syslog", "users",
]

_SHEET_NAMES = {
    "processes":  "Processes",
    "netconns":   "Network Connections",
    "auth":       "Authentication",
    "cmdhistory": "Command History",
    "users":      "Users",
    "files":      "Files",
    "cron":       "Cron Jobs",
    "services":   "Services",
    "rcscripts":  "RC Scripts",
    "syslog":     "Syslog",
}

_ARTIFACT_COLS: dict[str, list[str]] = {
    "processes":  ["id", "pid", "ppid", "uid", "user", "started", "command", "arguments"],
    "netconns":   ["id", "proto", "local_addr", "local_port", "remote_addr", "remote_port", "state", "pid"],
    "auth":       ["id", "uid", "username", "result", "time", "method", "source", "destination"],
    "cmdhistory": ["id", "user", "command", "time", "history_file", "file_index"],
    "users":      ["id", "username", "uid", "gid", "gecos", "home", "shell"],
    "files":      ["id", "path", "mode", "size", "uid", "gid", "md5", "inode", "atime", "mtime", "ctime", "crtime"],
    "cron":       ["id", "username", "command", "minute", "hour", "day_of_month", "month", "day_of_week", "source_type", "source_file", "source_file_modified"],
    "services":   ["id", "unit_name", "unit_type", "description", "exec_start", "exec_start_pre", "exec_stop", "run_user", "run_group", "service_type", "restart", "environment", "wanted_by", "source_file", "source_file_modified"],
    "rcscripts":  ["id", "path", "source_type", "run_context", "username", "shell", "interpreter", "file_size", "source_file_modified"],
    "syslog":     ["id", "timestamp", "hostname", "program", "pid", "severity", "event_type", "actor_user", "target_user", "source_ip", "command", "message", "source_file"],
}

_TIMELINE_COLS = [
    "timestamp", "type", "hostname", "ts_description",
    "summary", "source_ip", "dest_ip", "md5", "assoc_user", "notes",
]

# Wide columns that should get extra space
_WIDE_COLS = {"command", "arguments", "message", "description", "content", "path", "summary", "notes"}

# ── Styling helpers ───────────────────────────────────────────────────────────

_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
_HDR_FILL  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HDR_ALIGN = Alignment(horizontal="left", vertical="center")
_CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)


def _style_sheet(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font  = _HDR_FONT
        cell.fill  = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        width = 40 if header in _WIDE_COLS else max(12, min(len(header) + 6, 30))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


def _write_row(ws, values: list) -> None:
    ws.append(values)
    for col_idx in range(1, len(values) + 1):
        ws.cell(row=ws.max_row, column=col_idx).alignment = _CELL_ALIGN


# ── Value formatting ──────────────────────────────────────────────────────────

def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT%H:%M:%SZ")
    return str(val)


# ── Timeline event building ───────────────────────────────────────────────────

def _events_for_type(
    atype: str,
    objs: list,
    note_map: dict[tuple[str, int], str],
    hostname: str | None,
) -> list[dict]:
    rows = []
    for obj in objs:
        note = note_map.get((atype, obj.id), "")
        enrich = _enrich_event(atype, obj, hostname=hostname)

        if atype == "processes":
            rows.append({**enrich, "timestamp": obj.started, "type": atype,
                         "summary": _process_summary(obj), "notes": note})

        elif atype == "auth":
            rows.append({**enrich, "timestamp": obj.time, "type": atype,
                         "summary": _auth_summary(obj), "notes": note})

        elif atype == "cmdhistory":
            rows.append({**enrich, "timestamp": obj.time, "type": atype,
                         "summary": _cmdhistory_summary(obj), "notes": note})

        elif atype == "cron":
            user = f"{obj.username}: " if obj.username else ""
            rows.append({**enrich, "timestamp": obj.source_file_modified, "type": atype,
                         "summary": f"{user}{obj.command or ''}".strip(), "notes": note})

        elif atype == "services":
            label  = f"{obj.unit_name or ''}.{obj.unit_type or ''}".strip(".")
            detail = obj.description or obj.exec_start or ""
            rows.append({**enrich, "timestamp": obj.source_file_modified, "type": atype,
                         "summary": f"{label}: {detail}".rstrip(": "), "notes": note})

        elif atype == "rcscripts":
            user_part = f"{obj.username}: " if obj.username else ""
            rows.append({**enrich, "timestamp": obj.source_file_modified, "type": atype,
                         "summary": f"{user_part}{obj.path or ''}".strip(), "notes": note})

        elif atype == "files":
            for attr in ("mtime", "atime", "ctime", "crtime"):
                ts = getattr(obj, attr)
                if ts is None:
                    continue
                e = _enrich_event(atype, obj, file_attr=attr, hostname=hostname)
                rows.append({**e, "timestamp": ts, "type": atype,
                             "summary": obj.path or "", "notes": note})

        elif atype == "netconns":
            summary = (f"{obj.proto} {obj.local_addr}:{obj.local_port}"
                       f" → {obj.remote_addr}:{obj.remote_port} [{obj.state}]")
            rows.append({**enrich, "timestamp": None, "type": atype,
                         "summary": summary, "notes": note})

        elif atype == "syslog":
            rows.append({**enrich, "timestamp": obj.timestamp, "type": atype,
                         "summary": _syslog_summary(obj), "notes": note})

        elif atype == "users":
            summary = f"{obj.username or ''} (uid={obj.uid})"
            rows.append({**enrich, "timestamp": None, "type": atype,
                         "summary": summary, "notes": note})

    return rows


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_timeline_sheet(wb: openpyxl.Workbook, events: list[dict]) -> None:
    ws = wb.active
    ws.title = "Timeline"
    ws.append(_TIMELINE_COLS)
    _style_sheet(ws, _TIMELINE_COLS)
    for ev in events:
        _write_row(ws, [_fmt(ev.get(col)) for col in _TIMELINE_COLS])


def _build_artifact_sheet(
    wb: openpyxl.Workbook,
    atype: str,
    objs: list,
    note_map: dict[tuple[str, int], str],
) -> None:
    cols = _ARTIFACT_COLS.get(atype, [])
    headers = cols + ["notes"]
    ws = wb.create_sheet(title=_SHEET_NAMES.get(atype, atype))
    ws.append(headers)
    _style_sheet(ws, headers)
    for obj in objs:
        note = note_map.get((atype, obj.id), "")
        _write_row(ws, [_fmt(getattr(obj, col, None)) for col in cols] + [note])


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.get("/collections/{collection_id}/export")
def export_collection(
    collection_id: int,
    tag_ids: str = Query(..., description="Comma-separated tag IDs"),
    db: Session = Depends(get_db),
):
    col = db.query(UACCollection).filter_by(id=collection_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")

    ids = [int(x) for x in tag_ids.split(",") if x.strip().isdigit()]
    if not ids:
        raise HTTPException(status_code=422, detail="No valid tag IDs provided")

    # OR logic: any row tagged with any of the selected tags
    taggings = (
        db.query(Tagging)
        .filter(Tagging.collection_id == collection_id, Tagging.tag_id.in_(ids))
        .all()
    )

    grouped: dict[str, set[int]] = {}
    for t in taggings:
        grouped.setdefault(t.artifact_type, set()).add(t.artifact_id)

    if not grouped:
        raise HTTPException(status_code=404, detail="No tagged rows found for the selected tags")

    # Fetch notes for all tagged rows in one query per type
    note_map: dict[tuple[str, int], str] = {}
    for atype, artifact_ids in grouped.items():
        for n in db.query(Note).filter(
            Note.collection_id == collection_id,
            Note.artifact_type == atype,
            Note.artifact_id.in_(artifact_ids),
        ).all():
            note_map[(n.artifact_type, n.artifact_id)] = n.content

    # Fetch ORM objects in canonical order
    objs_by_type: dict[str, list] = {}
    for atype in _TYPE_ORDER:
        if atype not in grouped:
            continue
        model = _MODEL_MAP.get(atype)
        if model is None:
            continue
        objs_by_type[atype] = (
            db.query(model).filter(model.id.in_(grouped[atype])).all()
        )

    # Build timeline events (merged + sorted)
    all_events: list[dict] = []
    for atype, objs in objs_by_type.items():
        all_events.extend(_events_for_type(atype, objs, note_map, col.hostname))
    all_events.sort(key=lambda e: (e["timestamp"] is None, e["timestamp"] or datetime.min))

    # Build workbook
    wb = openpyxl.Workbook()
    _build_timeline_sheet(wb, all_events)
    for atype, objs in objs_by_type.items():
        _build_artifact_sheet(wb, atype, objs, note_map)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_host = (col.hostname or "collection").replace(" ", "_")
    filename   = f"{safe_host}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
